# -*- coding:utf-8 -*-
# !/usr/bin/env python
"""
@Project ：AutoTestingProject1
@File ：excel_comparator
@Time ：2025/8/8 9:16
@Author ：11031840
@Motto: 理解しあうのはとても大事なことです。理解とは误解の総体に过ぎないと言う人もいますし
@describe：
Excel文件比较工具

功能：
1. 比较两个Excel文件的差异，支持多种配置选项
2. 支持批量比较任务
3. 提供详细的差异报告

主要特性：
- 支持排除指定列
- 支持多列排序
- 支持指定比较列
- 处理可序列化数据
- 智能处理行顺序不一致问题
- 详细的差异报告输出
- 批量任务支持

@Project ：AutoTestingProject1
@File ：excel_comparator
@Time ：2025/8/8 9:16
@Author ：11031840
@Motto: 理解しあうのはとても大事なことです。理解とは误解の総体に过ぎないと言う人もいますし
"""
import os
import pandas as pd
import numpy as np
import json
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, colors
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.worksheet import Worksheet
import warnings
from typing import List, Dict, Tuple, Optional, Union, Set


class ExcelComparator:
    """Excel文件比较器"""

    # 定义常量
    RESULT_COLUMNS = [
        '3.0表名', '4.0表名', '对比结果',
        '不同点描述', '差异部分(3.0值)', '差异部分(4.0值)'
    ]

    # 样式定义
    HEADER_FILL = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")  # 金色
    GREEN_FILL = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # 绿色
    YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # 黄色
    RED_FILL = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")  # 红色
    GREEN_FONT = Font(color="000000")  # 绿色,用于
    RED_FONT = Font(color="000000")  # 红色

    def __init__(self):
        """初始化比较器"""
        self.primary_key = None
        self.exclude_columns = []
        self.sort_columns = []
        self.compare_columns = []
        self.df1 = None
        self.df2 = None
        self.df1_name = ""
        self.df2_name = ""
        self.has_error = False

    def read_excel(
            self,
            file1: str,
            file2: str,
            primary_key: str,
            exclude_columns: List[str] = None,
            sort_columns: List[str] = None,
            compare_columns: List[str] = None
    ) -> Tuple[pd.DataFrame, pd.DataFrame]:
        """
        读取两个Excel文件并进行预处理
        :param file1: 第一个Excel文件路径
        :param file2: 第二个Excel文件路径
        :param primary_key: 主键列名
        :param exclude_columns: 需要排除的列名列表
        :param sort_columns: 需要排序的列名列表
        :param compare_columns: 需要比较的列名列表
        :return: 处理后的两个DataFrame
        """
        # 初始化参数
        exclude_columns = exclude_columns or []
        sort_columns = sort_columns or []
        compare_columns = compare_columns or []

        self.df1_name = os.path.basename(file1)
        self.df2_name = os.path.basename(file2)
        self.primary_key = primary_key
        self.exclude_columns = exclude_columns
        self.sort_columns = sort_columns
        self.compare_columns = compare_columns
        self.has_error = False
        try:
            # 读取Excel文件
            self._read_excel_files(file1, file2)

            # 验证主键
            self._validate_primary_key()

            # 处理排除列
            self._process_exclude_columns()

            # 处理排序列
            self._process_sort_columns()

            # 处理比较列
            self._process_compare_columns()

        except Exception as e:
            self.has_error = True
            warnings.warn(f"文件预处理失败: {str(e)}")

        return self.df1, self.df2

    def _read_excel_files(self, file1: str, file2: str):
        """读取Excel文件并处理可能的异常"""
        try:
            self.df1 = pd.read_excel(file1)
            self.df2 = pd.read_excel(file2)

            # 确保数据是DataFrame格式
            if not isinstance(self.df1, pd.DataFrame) or not isinstance(self.df2, pd.DataFrame):
                raise ValueError("读取的文件不是有效的Excel数据格式")

        except Exception as e:
            raise ValueError(f"读取Excel文件失败: {str(e)}")

    def _validate_primary_key(self):
        """验证主键列是否存在"""
        if self.primary_key not in self.df1.columns:
            raise ValueError(f"主键 '{self.primary_key}' 在第一个文件中不存在")
        if self.primary_key not in self.df2.columns:
            raise ValueError(f"主键 '{self.primary_key}' 在第二个文件中不存在")

    def _process_exclude_columns(self):
        """处理排除列"""
        # 获取有效的排除列（在两个文件中都存在且不是主键）
        valid_exclude_cols = [
            col for col in self.exclude_columns
            if col in self.df1.columns and col in self.df2.columns and col != self.primary_key
        ]

        # 记录无效的排除列
        invalid_exclude_cols = set(self.exclude_columns) - set(valid_exclude_cols)
        if invalid_exclude_cols:
            warnings.warn(f"以下排除列无效（不存在或为主键）: {', '.join(invalid_exclude_cols)}")

        # 应用排除
        self.df1 = self.df1.drop(columns=invalid_exclude_cols, errors='ignore')
        self.df2 = self.df2.drop(columns=invalid_exclude_cols, errors='ignore')

        # 更新排除列列表
        self.exclude_columns = valid_exclude_cols

        # 更新其他参数中的列
        self._update_column_references(invalid_exclude_cols)

    def _update_column_references(self, invalid_columns: Set[str]):
        """更新其他参数中的列引用"""
        # 更新排序列
        self.sort_columns = [col for col in self.sort_columns if col not in invalid_columns]

        # 更新比较列
        self.compare_columns = [col for col in self.compare_columns if col not in invalid_columns]

    def _process_sort_columns(self):
        """处理排序列"""
        if not self.sort_columns:
            self.sort_columns = [self.primary_key]
            return
        # 确保主键在排序列中且是第一个
        if self.primary_key not in self.sort_columns:
            self.sort_columns.insert(0, self.primary_key)
            warnings.warn(f"已自动添加主键 '{self.primary_key}' 到排序列中")
        # 获取有效的排序列
        valid_sort_cols = [
            col for col in self.sort_columns
            if col in self.df1.columns and col in self.df2.columns
        ]

        # 记录无效的排序列
        invalid_sort_cols = set(self.sort_columns) - set(valid_sort_cols)
        if invalid_sort_cols:
            warnings.warn(f"以下排序列无效: {', '.join(invalid_sort_cols)}")

        # 应用排序
        if valid_sort_cols:
            self.df1 = self.df1.sort_values(by=valid_sort_cols).reset_index(drop=True)
            self.df2 = self.df2.sort_values(by=valid_sort_cols).reset_index(drop=True)

        # 更新排序列列表
        self.sort_columns = valid_sort_cols

    def _process_compare_columns(self):
        """处理比较列"""
        if not self.compare_columns:
            # 如果没有指定比较列，则比较所有列（排除已排除的列）
            self.compare_columns = [
                col for col in self.df1.columns
                if col not in self.exclude_columns and col in self.df2.columns
            ]
        else:
            # 确保主键在比较列中
            if self.primary_key not in self.compare_columns:
                self.compare_columns.insert(0, self.primary_key)
                warnings.warn(f"已自动添加主键 '{self.primary_key}' 到比较列中")

            # 获取有效的比较列
            valid_compare_cols = [
                col for col in self.compare_columns
                if col in self.df1.columns and col in self.df2.columns
            ]

            # 记录无效的比较列
            invalid_compare_cols = set(self.compare_columns) - set(valid_compare_cols)
            if invalid_compare_cols:
                warnings.warn(f"以下比较列无效: {', '.join(invalid_compare_cols)}")

            # 更新比较列列表
            self.compare_columns = valid_compare_cols

    def _create_error_result(self, error_message: str) -> pd.DataFrame:
        same_result = pd.DataFrame([{
            '3.0表名': self.df1_name,
            '4.0表名': self.df2_name,
            '对比结果': '异常',
            '不同点描述': error_message,
            '差异部分(3.0值)': None,
            '差异部分(4.0值)': None
        }])
        return same_result

    def compare(self) -> pd.DataFrame:
        """
        执行比较并返回差异结果
        :return: 包含比较结果的DataFrame
        """
        if self.has_error:
            return self._create_error_result("文件预处理失败，无法比较")

        if self.df1 is None or self.df2 is None:
            return self._create_error_result("文件读取失败，无法比较")

        # 初始化结果DataFrame
        result_df = pd.DataFrame(columns=self.RESULT_COLUMNS)

        # 检查行数差异
        len_df1 = len(self.df1)
        len_df2 = len(self.df2)

        if len_df1 != len_df2:
            # 行数不同，进行排除对比法
            result = self._compare_with_row_diff()
        else:
            # 行数相同，逐行对比
            result = self._compare_row_by_row()

        # 合并结果
        result_df = pd.concat([result_df, result], ignore_index=True)

        # 如果没有差异，添加一条"相同"记录
        if result_df.empty:
            same_result = pd.DataFrame([{
                '3.0表名': self.df1_name,
                '4.0表名': self.df2_name,
                '对比结果': '相同',
                '不同点描述': None,
                '差异部分(3.0值)': None,
                '差异部分(4.0值)': None
            }])
            result_df = pd.concat([result_df, same_result], ignore_index=True)

        return result_df

    def _compare_with_row_diff(self) -> pd.DataFrame:
        """处理行数不同的情况"""
        result_df = pd.DataFrame()

        # 获取主键值的列表（保留顺序和重复项）
        df1_keys = self.df1[self.primary_key].tolist()
        df2_keys = self.df2[self.primary_key].tolist()

        # 找出只在df1中存在的行
        only_in_df1 = [k for k in df1_keys if k not in df2_keys]
        if only_in_df1:
            temp_df = pd.DataFrame({
                '3.0表名': self.df1_name,
                '4.0表名': self.df2_name,
                '对比结果': '存在行数差异',
                '不同点描述': f"行只在 {self.df1_name} 中存在",
                '差异部分(3.0值)': list(only_in_df1),
                '差异部分(4.0值)': None
            })
            result_df = pd.concat([result_df, temp_df], ignore_index=True)

        # 找出只在df2中存在的行
        only_in_df2 = [k for k in df2_keys if k not in df1_keys]
        if only_in_df2:
            temp_df = pd.DataFrame({
                '3.0表名': self.df1_name,
                '4.0表名': self.df2_name,
                '对比结果': '存在行数差异',
                '不同点描述': f"行只在 {self.df2_name} 中存在",
                '差异部分(3.0值)': None,
                '差异部分(4.0值)': list(only_in_df2)
            })
            result_df = pd.concat([result_df, temp_df], ignore_index=True)

        # 对比两个表中都存在的行
        common_keys = [k for k in df1_keys if k in df2_keys]
        if common_keys:
            common_df1 = self.df1[self.df1[self.primary_key].isin(common_keys)]
            common_df2 = self.df2[self.df2[self.primary_key].isin(common_keys)]

            # 临时保存原始数据
            original_df1, original_df2 = self.df1, self.df2
            self.df1, self.df2 = common_df1, common_df2

            # 比较共同的行
            common_result = self._compare_row_by_row()
            result_df = pd.concat([result_df, common_result], ignore_index=True)

            # 恢复原始数据
            self.df1, self.df2 = original_df1, original_df2

        return result_df

    def _compare_row_by_row(self) -> pd.DataFrame:
        """逐行比较两个DataFrame"""
        result_df = pd.DataFrame()

        # 确保行数相同
        if len(self.df1) != len(self.df2):
            raise ValueError("两个DataFrame的行数不同，不能逐行比较")

        # 检查主键顺序是否一致
        if not np.array_equal(self.df1[self.primary_key].values, self.df2[self.primary_key].values):
            warnings.warn("主键顺序不一致，将按主键重新排序")
            self._align_dataframes_by_primary_key()

        # 收集差异行
        diff_rows = []
        for i in range(len(self.df1)):
            row_diff = self._compare_single_row(i)
            if row_diff:
                diff_rows.extend(row_diff)

        # 构建结果DataFrame
        if diff_rows:
            result_df = pd.DataFrame(diff_rows, columns=self.RESULT_COLUMNS)

        return result_df

    def _align_dataframes_by_primary_key(self):
        """按主键对齐两个DataFrame"""
        df1_sorted = self.df1.set_index(self.primary_key)
        df2_sorted = self.df2.set_index(self.primary_key)

        # 获取共同的主键并按df1的顺序排序
        common_keys = [k for k in self.df1[self.primary_key] if k in df2_sorted.index]
        self.df1 = df1_sorted.loc[common_keys].reset_index()
        self.df2 = df2_sorted.loc[common_keys].reset_index()

    def _compare_single_row(self, row_idx: int) -> List[Dict]:
        """比较单行数据"""
        row1 = self.df1.iloc[row_idx]
        row2 = self.df2.iloc[row_idx]

        # 检查主键是否匹配
        if row1[self.primary_key] != row2[self.primary_key]:
            warnings.warn(f"第{row_idx + 1}行的主键不匹配: {row1[self.primary_key]} vs {row2[self.primary_key]}")
            return []

        # 比较指定列
        diff_cols = []
        for col in self.compare_columns:
            if col == self.primary_key:
                continue

            val1 = row1[col]
            val2 = row2[col]

            if not self._compare_values(val1, val2):
                diff_cols.append(col)

        # 如果有差异，构建结果
        if diff_cols:
            return [{
                '3.0表名': self.df1_name,
                '4.0表名': self.df2_name,
                '对比结果': '不同',
                '不同点描述': f"{row_idx + 1}行 【3.0{row1[self.primary_key]}】【4.0{row2[self.primary_key]}】【{col}】",
                '差异部分(3.0值)': row1[col],
                '差异部分(4.0值)': row2[col]
            } for col in diff_cols]

        return []

    def _compare_values(self, val1, val2) -> bool:
        """
         比较两个值是否相等
        处理NaN值、可序列化数据以及普通比较
        :return: 如果值相等返回True，否则返回False
        """
        # 处理NaN值
        if pd.isna(val1) and pd.isna(val2):
            return True
        # 处理数值类型的微小差异（特别是浮点数）
        if isinstance(val1, (int, float)) and isinstance(val2, (int, float)):
            return abs(val1 - val2) < 1e-9

        # 处理字符串类型的空值比较
        if isinstance(val1, str) and isinstance(val2, str):
            if val1.strip() == "" and val2.strip() == "":
                return True
        # 处理可序列化数据
        if self._is_serializable(val1) and self._is_serializable(val2):
            try:
                # 尝试解析JSON字符串
                parsed_val1 = json.loads(val1) if isinstance(val1, str) else val1
                parsed_val2 = json.loads(val2) if isinstance(val2, str) else val2
                return parsed_val1 == parsed_val2
            except json.JSONDecodeError:
                pass

        # 普通比较
        return val1 == val2

    def _is_serializable(self, value) -> bool:
        """检查值是否可序列化"""
        try:
            json.dumps(value)
            return True
        except (TypeError, ValueError):
            return False

    def save_diff_result(self, result_df: pd.DataFrame, output_file: str):
        """
        保存差异结果到Excel文件

        参数:
            result_df: 包含比较结果的DataFrame
            output_file: 输出文件路径
        """
        # 创建Excel工作簿
        wb = Workbook()
        ws = wb.active

        # 写入表头
        ws.append(list(result_df.columns))

        # 设置表头样式
        for cell in ws[1]:
            cell.fill = self.HEADER_FILL

        # 写入数据
        for row in dataframe_to_rows(result_df, index=False, header=False):
            ws.append(row)

        # 合并相同结果的单元格
        self._merge_result_cells(ws, result_df)

        # 设置单元格样式
        self._apply_result_styles(ws, result_df)

        # 调整列宽
        self._adjust_column_width(ws)

        # 保存文件
        wb.save(output_file)

    def _merge_result_cells(self, ws: Worksheet, result_df: pd.DataFrame):
        """合并相同结果的单元格"""
        # 合并前三个列的相同值
        for col_idx in [1, 2, 3]:  # 3.0表名, 4.0表名, 对比结果
            current_value = None
            start_row = 2  # 数据从第2行开始
            merge_start = start_row

            for row_idx in range(start_row, len(result_df) + start_row + 1):
                cell_value = ws.cell(row=row_idx, column=col_idx).value if row_idx <= len(result_df) + 1 else None

                if cell_value != current_value:
                    if merge_start < row_idx - 1:
                        ws.merge_cells(
                            start_row=merge_start, start_column=col_idx,
                            end_row=row_idx - 1, end_column=col_idx
                        )
                    current_value = cell_value
                    merge_start = row_idx

    def _apply_result_styles(self, ws: Worksheet, result_df: pd.DataFrame):
        """应用结果样式"""
        # 获取对比结果列的索引
        result_col_idx = None
        for idx, col in enumerate(result_df.columns, 1):
            if col == '对比结果':
                result_col_idx = idx
                break

        if not result_col_idx:
            return

        # 应用样式
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            result_cell = row[result_col_idx - 1]  # 转换为0-based索引
            if result_cell.value == '相同':
                result_cell.font = self.GREEN_FONT
                result_cell.fill = self.GREEN_FILL
            elif result_cell.value == '不同':
                result_cell.font = self.RED_FONT
                result_cell.fill = self.YELLOW_FILL
            else:
                result_cell.font = ExcelComparator.RED_FONT
                result_cell.fill = ExcelComparator.RED_FILL

    def _adjust_column_width(self, ws: Worksheet):
        """自动调整列宽"""
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter

            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass

            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column_letter].width = adjusted_width


class ExcelBatchComparator:
    """Excel批量比较器"""

    def __init__(self, config_file: str = None):
        """初始化批量比较器"""
        self.tasks = []
        self.common_params = {
            'primary_key': 'ID',
            'exclude_columns': [],
            'sort_columns': [],
            'compare_columns': []
        }

        if config_file:
            self.load_tasks_from_excel(config_file)

    def set_common_params(self, primary_key: str, exclude_columns: List[str] = None,
                          sort_columns: List[str] = None, compare_columns: List[str] = None):
        """
        设置通用参数

        参数:
            primary_key: 主键列名
            exclude_columns: 排除列列表
            sort_columns: 排序列列表
            compare_columns: 比较列列表
        """
        self.common_params = {
            'primary_key': primary_key,
            'exclude_columns': exclude_columns or [],
            'sort_columns': sort_columns or [],
            'compare_columns': compare_columns or []
        }

    def load_tasks_from_excel(self, file_path: str):
        """
        从Excel加载对比任务表

        参数:
            file_path: 任务表文件路径
        """
        try:
            df = pd.read_excel(file_path)
            required_columns = ['待对比3.0表名', '待对比4.0表名']
            if not all(col in df.columns for col in required_columns):
                raise ValueError(f"任务表必须包含列: {', '.join(required_columns)}")

            for _, row in df.iterrows():
                self.add_task(
                    file1=row['待对比3.0表名'],
                    file2=row['待对比4.0表名'],
                    primary_key=row.get('主键', self.common_params['primary_key']),
                    exclude_columns=row.get('排除列', self.common_params['exclude_columns']),
                    sort_columns=row.get('排序列', self.common_params['sort_columns']),
                    compare_columns=row.get('对比列', self.common_params['compare_columns'])
                )
        except Exception as e:
            raise ValueError(f"加载任务表失败: {str(e)}")

    def add_task(self, file1: str, file2: str, **kwargs):
        """
        添加单个对比任务

        参数:
            file1: 第一个Excel文件路径
            file2: 第二个Excel文件路径
            kwargs: 其他参数(primary_key, exclude_columns, sort_columns, compare_columns)
        """
        task = {
            'file1': file1,
            'file2': file2,
            'primary_key': kwargs.get('primary_key', self.common_params['primary_key']),
            'exclude_columns': kwargs.get('exclude_columns', self.common_params['exclude_columns']),
            'sort_columns': kwargs.get('sort_columns', self.common_params['sort_columns']),
            'compare_columns': kwargs.get('compare_columns', self.common_params['compare_columns'])
        }
        self.tasks.append(task)

    def run_all_tasks(self, output_file: str = "diff_result.xlsx",
                      base_path_3: str = "", base_path_4: str = ""):
        """
        执行所有对比任务
        :param output_file: 输出文件路径
        :param base_path_3: 3.0文件基础路径
        :param base_path_4: 4.0文件基础路径
        """
        if not self.tasks:
            print("没有可执行的对比任务")
            return

        # 初始化结果DataFrame
        all_results = pd.DataFrame()

        for i, task in enumerate(self.tasks, 1):
            try:
                # 构建完整文件路径
                file1 = os.path.join(base_path_3, task['file1'])
                file2 = os.path.join(base_path_4, task['file2'])
                if not os.path.exists(file1):
                    raise FileNotFoundError(f"3.0文件不存在: {file1}")
                if not os.path.exists(file2):
                    raise FileNotFoundError(f"4.0文件不存在: {file2}")
                print(f"\n正在执行任务 {i}/{len(self.tasks)}: {task['file1']} vs {task['file2']}")

                # 创建比较器实例
                comparator = ExcelComparator()

                # 读取并预处理Excel文件
                df1, df2 = comparator.read_excel(
                    file1, file2, task['primary_key'],
                    task['exclude_columns'], task['sort_columns'], task['compare_columns']
                )
                # 检查列是否匹配
                if not set(df1.columns).issuperset(set(task['compare_columns'] or df1.columns)):
                    missing_cols = set(task['compare_columns'] or df1.columns) - set(df1.columns)
                    warnings.warn(f"3.0文件中缺少列: {', '.join(missing_cols)}")

                if not set(df2.columns).issuperset(set(task['compare_columns'] or df2.columns)):
                    missing_cols = set(task['compare_columns'] or df2.columns) - set(df2.columns)
                    warnings.warn(f"4.0文件中缺少列: {', '.join(missing_cols)}")
                # 执行比较
                diff_result = comparator.compare()

                # 添加任务信息
                diff_result['任务编号'] = i
                diff_result['3.0文件'] = task['file1']
                diff_result['4.0文件'] = task['file2']

                # 重新排列列顺序
                cols = ['任务编号', '3.0文件', '4.0文件'] + [
                    c for c in diff_result.columns if c not in ['任务编号', '3.0文件', '4.0文件']
                ]
                diff_result = diff_result[cols]

                # 合并到总结果
                all_results = pd.concat([all_results, diff_result], ignore_index=True)

                print(f"对比完成，发现差异: {len(diff_result[diff_result['对比结果'] == '不同'])}条")

            except Exception as e:
                print(f"对比失败: {str(e)}")
                # 记录失败任务
                failed_result = pd.DataFrame([{
                    '任务编号': i,
                    '3.0文件': task['file1'],
                    '4.0文件': task['file2'],
                    '3.0表名': 'N/A',
                    '4.0表名': 'N/A',
                    '对比结果': '失败',
                    '不同点描述': str(e),
                    '差异部分(3.0值)': 'N/A',
                    '差异部分(4.0值)': 'N/A'
                }])
                all_results = pd.concat([all_results, failed_result], ignore_index=True)

        # 保存到Excel文件
        if not all_results.empty:
            self._save_to_excel(all_results, output_file)
            print(f"\n所有任务完成，结果已保存到: {output_file}")
        else:
            print("\n没有生成任何结果")

    def _save_to_excel(self, df: pd.DataFrame, output_file: str):
        """保存结果到Excel文件"""
        # 创建Excel工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "对比结果"

        # 写入表头
        headers = list(df.columns)
        ws.append(headers)

        # 设置表头样式
        for cell in ws[1]:
            cell.fill = ExcelComparator.HEADER_FILL

        # 写入数据
        for row in dataframe_to_rows(df, index=False, header=False):
            ws.append(row)

        # 设置样式
        self._apply_styles(ws, df)

        # 合并任务单元格
        self._merge_task_cells(ws, df)

        # 调整列宽
        self._adjust_column_width(ws)

        # 保存文件
        wb.save(output_file)

    def _apply_styles(self, ws: Worksheet, df: pd.DataFrame):
        """应用单元格样式"""
        # 获取对比结果列的索引
        result_col_idx = None
        for idx, col in enumerate(df.columns, 1):
            if col == '对比结果':
                result_col_idx = idx
                break

        # 应用样式
        if result_col_idx:
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                result_cell = row[result_col_idx - 1]  # 转换为0-based索引
                if result_cell.value == '相同':
                    result_cell.font = ExcelComparator.GREEN_FONT
                    result_cell.fill = ExcelComparator.GREEN_FILL
                elif result_cell.value == '不同':
                    result_cell.font = ExcelComparator.RED_FONT
                    result_cell.fill = ExcelComparator.YELLOW_FILL
                else:
                    result_cell.font = ExcelComparator.RED_FONT
                    result_cell.fill = ExcelComparator.RED_FILL

    def _merge_task_cells(self, ws: Worksheet, df: pd.DataFrame):
        """合并相同任务的前三列单元格"""
        # 获取任务编号列的索引
        task_col_idx = None
        for idx, col in enumerate(df.columns, 1):
            if col == '任务编号':
                task_col_idx = idx
                break

        if not task_col_idx:
            return

        # 合并前三列（任务编号、3.0文件、4.0文件）
        for col_idx in [task_col_idx, task_col_idx + 1, task_col_idx + 2]:
            current_value = None
            start_row = 2  # 数据从第2行开始
            merge_start = start_row

            for row_idx in range(start_row, len(df) + start_row + 1):
                cell_value = ws.cell(row=row_idx, column=col_idx).value if row_idx <= len(df) + 1 else None

                if cell_value != current_value:
                    if merge_start < row_idx - 1:
                        ws.merge_cells(
                            start_row=merge_start, start_column=col_idx,
                            end_row=row_idx - 1, end_column=col_idx
                        )
                    current_value = cell_value
                    merge_start = row_idx

    def _adjust_column_width(self, ws: Worksheet):
        """自动调整列宽"""
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter

            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass

            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column_letter].width = adjusted_width


if __name__ == "__main__":
    # 示例用法
    try:
        # 创建比较器实例
        comparator = ExcelComparator()

        # 定义参数
        file1 = r"D:\data\data_4\J0948012400000-1Y.xlsx"
        file2 = r"D:\data\data_3\J0948012400000-1Y.xlsx"
        primary_key = "名称"  # 主键列名
        exclude_columns = []  # 要排除的列
        sort_columns = ["物料编码", "名称", "数量", "长", "高"]  # 排序依据的列
        compare_columns = []  # 要比较的列

        # 读取并预处理Excel文件
        df1, df2 = comparator.read_excel(
            file1, file2, primary_key,
            exclude_columns, sort_columns, compare_columns
        )

        # 执行比较
        diff_result = comparator.compare()

        # 保存差异结果
        output_file = r"D:\data\diff_result.xlsx"
        comparator.save_diff_result(diff_result, output_file)

        print(f"比较完成，结果已保存到 {output_file}")


    except Exception as e:
        print(f"比较过程中发生错误: {str(e)}")
    # 批量比较示例
    # batch_comparator = ExcelBatchComparator()
    # batch_comparator.set_common_params(primary_key="ID")
    # batch_comparator.add_task("file1.xlsx", "file2.xlsx")
    # batch_comparator.add_task("file3.xlsx", "file4.xlsx")
    # batch_comparator.run_all_tasks("batch_diff_result.xlsx")
